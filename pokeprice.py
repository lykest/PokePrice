#!/usr/bin/env python3
import requests
import sys
from bs4 import BeautifulSoup
import time
import openpyxl
import os

'''
PokePrice -- A utility to take an Excel spreadsheet (database) full of monsters and figure out what their approximate value 
is from scraping a price aggregator with feeds from several auction sites.

The program expects a single-sheet Excel workbook with the first row as labels for the following five fields (columns):

    Name  <Req> |  Type <Opt> |  Marking <Opt> |  EN count <Opt> |  JP count <Opt>

    ex:  Pikachu  |  Electric  |  PROMO  |  2  |  5

The EN and JP count are used as multipliers for the determined average, low, and high values for a given Name.
Only Name is required for practical functionality, as the program assumes EN as 1 if not specified.

Extraneous columns are ignored by the program.  Only the Name (first) column is required.

'''


SITE_SEARCH='https://mavin.io/search?q={}&bt=sold'
#
# You may want to change this.
#
OUTPUT_FILENAME='Output.xlsx'
path_to_file = 'pokemon_cards.xlsx'
##############################

total_lowest_price = 0.0
total_highest_price = 0.0
total_average_price = 0.0

if os.path.exists(OUTPUT_FILENAME):
    print('Please move or delete the output file {}, we don\'t want to overwrite it'.format(OUTPUT_FILENAME))
    sys.exit(1)

result_workbook = openpyxl.Workbook()
result_sheet = result_workbook.active
result_sheet.append(('Name', 'Lowest Value', 'Highest Value', 'Average Value', 'Calc Lowest Value', 'Calc Highest Value', 'Calc Average Value'))


def write_totals():
    result_sheet.append(())
    result_sheet.append(('Total Average Price', 'Total Lowest Price', 'Total Highest Price'))
    result_sheet.append((total_average_price, total_lowest_price, total_highest_price))


def save_results(error = 0):
    write_totals()
    result_workbook.save(OUTPUT_FILENAME)

    sys.exit(error)



class Card:
    '''
    Defines an entity representing a Pokemon card.
    '''
    name = ''
    lowest_value = 0.0
    average_value = 0.0
    highest_value = 0.0
    count_english = 0
    count_japanese = 0


    def __init__(self, name, en_count = 0, jp_count = 0):
        self.name = name
        self.count_english = en_count
        self.count_japanese = jp_count

        if en_count == 0 and jp_count == 0:
            self.count_english = 1

    def write_value(self):
        """
        This function will calculate the relevant values based on card count and append them to a new Workbook.
        """
        calc_lowest_value = self.lowest_value * (self.count_english + self.count_japanese)
        calc_highest_value = self.highest_value * (self.count_english + self.count_japanese)
        calc_average_value = self.average_value * (self.count_english + self.count_japanese)

        # Update our global counters
        global total_lowest_price
        global total_highest_price
        global total_average_price
        total_lowest_price += calc_lowest_value
        total_highest_price += calc_highest_value
        total_average_price += calc_average_value

        result_sheet.append((self.name, self.lowest_value, self.highest_value, self.average_value, calc_lowest_value, calc_highest_value, calc_average_value))

    def set_value(self):
        """
        This function will search the Mavin database and find the current value for one card.
        """
        lookup_page = requests.get(SITE_SEARCH.format(self.name))
        if lookup_page.status_code != 200:
            # We abort here
            save_results()

        soup = BeautifulSoup(lookup_page.content, 'html.parser')
        # select('#id') also works
        average_price = soup.find(id='medianHiddenField')
        lowest_price = soup.find(id='lowestWorthItemHeader')
        highest_price = soup.find(id='highestWorthItemHeader')

        if not average_price or not lowest_price or not highest_price:
            """
            It turns out sometimes there are no results.  In such a case, we should continue.
            """
            self.average_value = 0.0
            self.lowest_value = 0.0
            self.highest_value = 0.0

        else:
            try:
                self.average_value = float(average_price['value'])
                # Have to chomp the dollar ($) off the front.
                # Apparently this does not always work.
                # Replace manually.
                self.lowest_value = float(lowest_price['data-sold'][1:].replace('$', '').replace(',', ''))
                self.highest_value = float(highest_price['data-sold'][1:].replace('$', '').replace(',', ''))
            except Exception as e:
                print('Error extracting the values from the elements')
                save_results()
                # raise(e)

        return


    def get_highest_value(self):
        return self.highest_value

    def get_lowest_value(self):
        return self.lowest_value

    def get_average_value(self):
        return self.average_value

    def get_name(self):
        return self.name



def parse_workbook(file_hdl):
    return

def add_fields(file_hdl):
    return


list_of_cards = []

workbook = openpyxl.load_workbook(path_to_file)
sheet = workbook.active

max_column = sheet.max_column
max_row = sheet.max_row

# Skip the first row, as these are all labels.
for i in range(2, max_row + 1):
    poke_name = sheet.cell(row = i, column = 1).value
    poke_type = sheet.cell(row = i, column = 2).value
    poke_marking = sheet.cell(row = i, column = 3).value
    poke_en_count = int(sheet.cell(row = i, column = 4).value or 0)
    poke_jp_count = int(sheet.cell(row = i, column = 5).value or 0)

    list_of_cards.append(Card(poke_name, poke_en_count, poke_jp_count))

for card in list_of_cards:
    card.set_value()
    card.write_value()
    time.sleep(2)          # Don't abuse the service. 

save_results()
